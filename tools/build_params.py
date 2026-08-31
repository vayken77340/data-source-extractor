"""Turn a business spreadsheet into a parameter file.

Run by hand when the business sends a new sheet, then commit the result. This is the only
thing in the repo that knows what a merged cell is — everything downstream reads the one
shape written here, so a second spreadsheet with different columns needs a different
invocation of this script, not a different provider.

    python tools/build_params.py input/asset_types.xlsx \
        --sheet Referentiel --columns assetType,measureType --ffill assetType \
        -o config/params/asset_types.json

A column may be renamed on the way out with `header=name`, which is how a sheet written in
the business's vocabulary produces a parameter file written in yours:

    --columns "Type d'actif=assetType,Grandeur mesuree=measureType" --ffill assetType

Deliberately not in `providers/`: every `.py` there is imported at startup, and a build
script has no business running during `validate` or `list-providers`.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SCHEMA = "param-file/1"


def clean(cell: Any) -> Any:
    """Blank means blank: whitespace-only strings are nothing."""
    if isinstance(cell, str):
        return cell.strip() or None
    return cell


def read_sheet(workbook_path: Path, sheet: str) -> list[tuple[Any, ...]]:
    if not workbook_path.is_file():
        raise SystemExit(f"no such file: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise SystemExit(
                f"{workbook_path} has no sheet {sheet!r} "
                f"(sheets: {', '.join(workbook.sheetnames)})"
            )
        return list(workbook[sheet].iter_rows(values_only=True))
    finally:
        workbook.close()


def parse_columns(specs: list[str]) -> dict[str, str]:
    """`header=name` pairs, or a bare `header` meaning no rename. Sheet order is kept.

    Renaming belongs here rather than downstream. The sheet's vocabulary belongs to the
    business and changes when they feel like it; the parameter file's belongs to you and is
    what YAML markers, `bind` keys and output templates are written against. Past this
    point nothing needs to know that `assetType` was once `Type d'actif`.

    Deliberately explicit rather than inferred — camel-casing a header is a guess, and a
    guess that breaks when someone adds an accent is worse than a mapping you can read.
    """
    columns: dict[str, str] = {}
    for spec in specs:
        header, _, renamed = spec.partition("=")
        header, name = header.strip(), (renamed.strip() or header.strip())
        if not header:
            raise SystemExit(f"empty column in --columns: {spec!r}")
        if header in columns:
            raise SystemExit(f"--columns names the sheet column {header!r} twice")
        if name in columns.values():
            raise SystemExit(f"--columns maps two sheet columns to {name!r}")
        columns[header] = name
    if not columns:
        raise SystemExit("--columns is empty")
    return columns


def extract(
    rows: list[tuple[Any, ...]], columns: dict[str, str], ffill: frozenset[str]
) -> list[dict[str, Any]]:
    """Project `columns` (sheet header -> output name), forward-filling `ffill`, dropping
    blanks, deduping. `ffill` names output columns, since everything past the header row is
    in your vocabulary rather than the sheet's.

    Forward-fill is how a merged cell survives. In read-only mode openpyxl returns None for
    every cell of a merged range but the top-left, and does not populate `merged_cells` at
    all — so a blank in a forward-filled column means "the same value as the row above",
    which is exactly what the merge means on screen.

    Declared per column rather than inferred, because a blank meaning "missing" and a blank
    meaning "same as above" are indistinguishable from here. Only you know which sheet is
    which.
    """
    if not rows:
        return []

    header = [clean(cell) for cell in rows[0]]
    missing = [source for source in columns if source not in header]
    if missing:
        present = ", ".join(str(cell) for cell in header if cell is not None)
        raise SystemExit(f"no column(s) {missing} in the sheet (headers: {present})")
    indexes = {name: header.index(source) for source, name in columns.items()}
    names = list(columns.values())

    carried: dict[str, Any] = {}
    out: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for row in rows[1:]:
        record: dict[str, Any] = {}
        for name, index in indexes.items():
            value = clean(row[index]) if index < len(row) else None
            if name in ffill:
                # `.get`, not `[]`: a sheet whose very first data row is blank here has
                # nothing to carry down, and that row should drop rather than crash.
                value = carried.get(name) if value is None else value
                carried[name] = value
            record[name] = value

        cells = tuple(record[name] for name in names)
        if any(cell is None for cell in cells) or cells in seen:
            continue
        seen.add(cells)
        out.append(record)
    return out


def build(workbook_path: Path, sheet: str, columns: list[str], rows: list[dict[str, Any]]) -> dict:
    """Provenance plus rows. The provenance is half the point: it is what tells you, six
    weeks later, which spreadsheet a given output tree came from."""
    return {
        "schema": SCHEMA,
        "source_file": workbook_path.as_posix(),
        "sheet": sheet,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "columns": columns,
        "rows": rows,
    }


def write(path: Path, document: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(document, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def summarize(rows: list[dict[str, Any]], columns: list[str]) -> str:
    """Counts, so a sheet that quietly halved is visible without opening the file."""
    counts = ", ".join(
        f"{column} {len({row[column] for row in rows})} distinct" for column in columns
    )
    return f"{len(rows)} rows ({counts})"


def split(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("workbook", type=Path, help="the spreadsheet the business supplied")
    parser.add_argument("-o", "--out", type=Path, required=True, help="parameter file to write")
    parser.add_argument("--sheet", required=True)
    parser.add_argument(
        "--columns",
        required=True,
        action="append",
        help="comma-separated, in output order; may also be repeated. `header=name` "
        "renames a column on the way out, a bare `header` keeps it. A sheet header "
        "containing a comma or an `=` cannot be named — rename it in the sheet.",
    )
    parser.add_argument(
        "--ffill",
        default="",
        help="comma-separated subset of the *output* names whose blanks mean 'same as the "
        "row above' — i.e. the merged ones",
    )
    args = parser.parse_args(argv)

    columns = parse_columns([spec for group in args.columns for spec in split(group)])
    names = list(columns.values())
    ffill = frozenset(split(args.ffill))
    unknown = sorted(ffill - set(names))
    if unknown:
        raise SystemExit(f"--ffill names {unknown}, which is not in --columns {names}")

    rows = extract(read_sheet(args.workbook, args.sheet), columns, ffill)
    if not rows:
        raise SystemExit(
            f"{args.workbook}[{args.sheet}] yielded no rows — wrong sheet or columns?"
        )

    write(args.out, build(args.workbook, args.sheet, names, rows))
    print(f"{summarize(rows, names)} -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
