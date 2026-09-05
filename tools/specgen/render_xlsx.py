"""Project the model into the companion workbook with openpyxl.

Sheet names, column headers and every cell label come from the label file; the sheet list
comes from the model (`appendix.workbook.tabs`), which the .docx also renders in its
annex — so the two cannot disagree. One sheet per endpoint, named after the endpoint,
holds the structure of its response as observed. Nothing is computed here.
"""

from __future__ import annotations

from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from specgen import labels
from specgen.labels import L

HEADER_FILL = PatternFill("solid", fgColor="EDEDED")
WIDTH_MAX = 70


def render(model: Mapping[str, Any], out: Path) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    by_name = {endpoint["name"]: endpoint for endpoint in model["endpoints"]}
    by_key = {entry["key"]: entry for entry in model["appendix"]["workbook"]["lists"]}
    for tab in model["appendix"]["workbook"]["tabs"]:
        sheet = workbook.create_sheet(tab["name"])
        key = tab["key"]
        if key == "readme":
            _readme(sheet, model)
        elif key == "endpoints":
            _endpoints(sheet, model)
        elif key == "metadata":
            _metadata(sheet, model)
        elif key.startswith("list:"):
            _value_list(sheet, by_key[key])
        elif key.startswith("response:"):
            _endpoint_sheet(sheet, by_name[key.split(":", 1)[1]])
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)
    return out


def _value_list(sheet, entry: Mapping[str, Any]) -> None:
    """A parameter list's actual values.

    This is why the referential no longer travels as a JSON file: it is rows and columns,
    the workbook is already being delivered, and a spreadsheet is where a table is read.
    """
    start = 1
    if entry["generated"]:
        sheet.cell(row=1, column=1, value=L.fmt("workbook.list.generated", date=entry["generated"])).font = Font(italic=True)
        start = 3
    _table(sheet, list(entry["columns"]), entry["rows"], start_row=start)
    if not entry["rows"]:
        sheet.cell(row=start + 1, column=1, value=str(L["workbook.list.empty"])).font = Font(italic=True)


def _table(sheet, headers: Sequence[str], rows: Sequence[Sequence[Any]], start_row: int = 1) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=start_row, column=column, value=header)
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for offset, row in enumerate(rows, start=1):
        for column, value in enumerate(row, start=1):
            cell = sheet.cell(row=start_row + offset, column=column, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = sheet.cell(row=start_row + 1, column=1)
    for column in range(1, len(headers) + 1):
        longest = max(
            (len(str(sheet.cell(row=r, column=column).value or "")) for r in range(start_row, start_row + len(rows) + 1)),
            default=10,
        )
        current = sheet.column_dimensions[get_column_letter(column)].width or 0
        sheet.column_dimensions[get_column_letter(column)].width = max(current, min(max(12, longest + 2), WIDTH_MAX))


def _readme(sheet, model: Mapping[str, Any]) -> None:
    document = model["document"]
    done = model["completeness"]
    rows = L.section("workbook.readme.rows")
    intro = [
        (rows["document"], document["title"]),
        (rows["source_system"], document["source_system"]),
        (rows["version"], document["version_text"]),
        (rows["date"], document["date"]),
        (rows["generated"], labels.fr_date(model["generated_at"])),
        (
            rows["progress"],
            str(rows["progress_value"]).format(
                percent=labels.fr_decimal(done["percent"]),
                count=labels.plural(done["todo"], str(rows["progress_unit"])),
            ),
        ),
        (rows["convention"], rows["convention_value"]),
    ]
    if model["links"]["document"]:
        intro.insert(1, (rows["document_link"], document["file"]))
    _table(sheet, list(L["workbook.readme.columns"]), intro)
    if model["links"]["document"]:
        cell = sheet.cell(row=3, column=2)  # the row just inserted, under the header
        cell.hyperlink = model["links"]["document"]
        cell.font = Font(color="0563C1", underline="single")
    start = len(intro) + 3
    sheet.cell(row=start, column=1, value=str(L["workbook.readme.tabs_title"])).font = Font(bold=True)
    tabs = [(tab["name"], tab["contents"], tab["reader"]) for tab in model["appendix"]["workbook"]["tabs"]]
    _table(sheet, list(L["workbook.readme.tabs_columns"]), tabs, start_row=start + 1)


def _endpoints(sheet, model: Mapping[str, Any]) -> None:
    columns = L.section("workbook.endpoints.columns")
    rows = []
    for endpoint in model["endpoints"]:
        pagination = "; ".join(
            L.fmt("workbook.endpoints.pagination_cell", item=row["item"], value=row["value"])
            for row in (endpoint["pagination"] or [])
        )
        params = "; ".join(
            L.fmt("workbook.endpoints.param_cell", name=p["name"], location=p["location"], origin=p["origin"])
            for p in endpoint["params"]
        )
        rows.append(
            {
                "name": endpoint["name"],
                "method": endpoint["method"],
                "path": endpoint["path"],
                "purpose": _summary_value(endpoint, L["endpoint.summary.purpose"]),
                "depends_on": ", ".join(endpoint["depends_on"]),
                "params": params,
                "pagination": pagination,
                "files": endpoint["files"],
                "mode": endpoint["mode"],
                "grain": _summary_value(endpoint, L["endpoint.summary.grain"]),
                "key": endpoint["rendered_key"],
            }
        )
    # The label file decides which columns exist and in what order.
    _table(sheet, list(columns.values()), [[row[key] for key in columns] for row in rows])


def _summary_value(endpoint: Mapping[str, Any], item: str) -> str:
    return next((row["value"] for row in endpoint["summary"] if row["item"] == item), "")


def _metadata(sheet, model: Mapping[str, Any]) -> None:
    rows = [(a["attribute"], a["type"], a["mandatory"], a["description"]) for a in model["landing"]["contract"]]
    _table(sheet, list(L["workbook.metadata.columns"]), rows)


def _endpoint_sheet(sheet, endpoint: Mapping[str, Any]) -> None:
    """One endpoint, both directions: what it sends above what it returns.

    The document shows the request's *shape*; this is the field-by-field reference, and
    putting it beside the response means one sheet answers everything about an endpoint.
    """
    sheet.cell(row=1, column=1, value=str(L["workbook.blocks.request"])).font = Font(bold=True)
    end = _request(sheet, endpoint, start_row=2)
    sheet.cell(row=end + 2, column=1, value=str(L["workbook.blocks.response"])).font = Font(bold=True)
    _response(sheet, endpoint, start_row=end + 3)


def _request(sheet, endpoint: Mapping[str, Any], start_row: int) -> int:
    columns = L.section("workbook.request.columns")
    rows = [[p[key] for key in columns] for p in endpoint["params"]]
    _table(sheet, list(columns.values()), rows, start_row=start_row)
    if not rows:
        sheet.cell(row=start_row + 1, column=1, value=str(L["workbook.request.empty"])).font = Font(italic=True)
        return start_row + 1
    return start_row + len(rows)


def _response(sheet, endpoint: Mapping[str, Any], start_row: int = 1) -> int:
    columns = L.section("workbook.response.columns")
    fields = endpoint["response_fields"]
    if not fields:
        _table(sheet, list(columns.values()), [], start_row=start_row)
        sheet.cell(row=start_row + 1, column=1, value=str(L["workbook.response.empty"])).font = Font(italic=True)
        return start_row + 1
    rows = [
        {
            "path": row["path"],
            "type": row["type"],
            "nullable": labels.yes_no(row["nullable"]),
            "presence": row["presence"],
            "example": row["example"],
        }
        for row in fields
    ]
    _table(sheet, list(columns.values()), [[row[key] for key in columns] for row in rows], start_row=start_row)
    if "presence" in columns:
        column = list(columns).index("presence") + 1
        for r in range(start_row + 1, start_row + len(rows) + 1):
            sheet.cell(row=r, column=column).number_format = "0 %"
    return start_row + len(rows)
