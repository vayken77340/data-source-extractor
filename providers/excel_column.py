"""Read parameter values out of a spreadsheet the business supplies."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from api_extractor.providers import ProviderContext, provider


@provider("excel_column")
def excel_column(
    ctx: ProviderContext, *, path: str, sheet: str, columns: list[str]
) -> list[dict[str, Any]]:
    """One or more columns off a sheet, keyed by header name.

    Several columns stay row-wise — one spreadsheet row in, one dict out — so `assetType`
    and `region` off the same row remain correlated. A row is dropped when any requested
    cell is blank, since a blank parameter is not worth a request. Duplicate rows collapse,
    first occurrence wins, order is otherwise preserved.
    """
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"excel_column: no such file: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet not in workbook.sheetnames:
            raise ValueError(
                f"excel_column: {workbook_path} has no sheet {sheet!r} "
                f"(sheets: {', '.join(workbook.sheetnames)})"
            )
        rows = list(workbook[sheet].iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    header = [_clean(cell) for cell in rows[0]]
    missing = [column for column in columns if column not in header]
    if missing:
        present = ", ".join(str(cell) for cell in header if cell is not None)
        raise ValueError(
            f"excel_column: {workbook_path}[{sheet}] has no column(s) {missing} "
            f"(headers: {present})"
        )
    indexes = [header.index(column) for column in columns]

    out: list[dict[str, Any]] = []
    seen: set[tuple[Any, ...]] = set()
    for row in rows[1:]:
        cells = tuple(_clean(row[i]) if i < len(row) else None for i in indexes)
        if any(cell is None for cell in cells) or cells in seen:
            continue
        seen.add(cells)
        out.append(dict(zip(columns, cells, strict=True)))
    return out


def _clean(cell: Any) -> Any:
    """Blank means blank: whitespace-only strings are nothing."""
    if isinstance(cell, str):
        return cell.strip() or None
    return cell
