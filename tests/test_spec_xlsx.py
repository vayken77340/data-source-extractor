"""The workbook is a projection of the model: sheets from the model, labels from the file,
numbers as numbers, and one sheet per endpoint holding its response's structure."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from api_extractor.providers.registry import SavedOutput
from specgen import evidence as ev, model, render_xlsx
from specgen.labels import L
from tests.conftest import build_envelope


@pytest.fixture
def workbook(reference_source, reference_annotation, tmp_path):
    built = model.build(reference_source, reference_annotation)
    path = render_xlsx.render(built.model, tmp_path / "annexes.xlsx")
    return built.model, load_workbook(path)


def column(sheet, index: int, start: int = 2) -> list:
    return [sheet.cell(row=r, column=index).value for r in range(start, sheet.max_row + 1)]


def header(sheet, row: int = 1) -> list:
    return [sheet.cell(row=row, column=c).value for c in range(1, sheet.max_column + 1) if sheet.cell(row=row, column=c).value]


def block(sheet, title: str) -> int:
    """The header row of a titled block — an endpoint sheet holds request then response."""
    for r in range(1, sheet.max_row + 1):
        if sheet.cell(row=r, column=1).value == title:
            return r + 1
    raise AssertionError(f"no block titled {title!r} in {sheet.title}")


def test_sheets_are_the_fixed_three_then_the_lists_then_the_endpoints(workbook):
    built, book = workbook
    assert book.sheetnames == [tab["name"] for tab in built["appendix"]["workbook"]["tabs"]]
    assert book.sheetnames == [
        "Readme", "Endpoints", "Metadata",
        "types d'actifs",
        "assets", "alarms", "tenant_info", "measures",
    ]


def test_a_value_list_ships_its_rows_rather_than_a_json_file(workbook):
    """The referential is rows and columns; the workbook is already being delivered."""
    _built, book = workbook
    sheet = book["types d'actifs"]
    assert header(sheet) == ["assetType"]
    assert column(sheet, 1) == ["PUMP", "VALVE"]


def test_a_generated_referential_carries_its_date(reference_source, reference_annotation, tmp_path, monkeypatch):
    """A referential with no date is one nobody can tell is stale."""
    import json as _json

    referential = tmp_path / "config" / "params" / "kinds.json"
    referential.parent.mkdir(parents=True)
    referential.write_text(
        _json.dumps({"generated_at": "2026-08-31T14:11:26Z", "columns": ["kind"], "rows": [{"kind": "PUMP"}]}),
        encoding="utf-8",
    )
    monkeypatch.chdir(tmp_path)
    source = reference_source.model_copy(deep=True)
    source.providers["asset_types"].args = {"path": "config/params/kinds.json", "columns": ["kind"]}
    source.providers["asset_types"].fn = "param_file"
    built = model.build(source, reference_annotation)
    entry = built.model["appendix"]["workbook"]["lists"][0]
    assert entry["generated"] == "31/08/2026"
    book = load_workbook(render_xlsx.render(built.model, tmp_path / "a.xlsx"))
    assert book[entry["sheet"]].cell(row=1, column=1).value == "Référentiel généré le 31/08/2026"


def test_the_endpoints_sheet_has_the_configured_columns_and_no_more(workbook):
    built, book = workbook
    sheet = book["Endpoints"]
    assert header(sheet) == list(L.section("workbook.endpoints.columns").values())
    assert column(sheet, 1) == [e["name"] for e in built["endpoints"]]
    for gone in ("Appelé", "N°"):
        assert gone not in header(sheet)
    assert sheet.freeze_panes == "A2"


def test_the_metadata_sheet_uses_iceberg_types(workbook):
    built, book = workbook
    sheet = book["Metadata"]
    assert column(sheet, 1) == [row["attribute"] for row in built["landing"]["contract"]]
    assert set(column(sheet, 2)) <= set(L.section("types.contract").values())
    assert "timestamp" in column(sheet, 2) and "int" in column(sheet, 2)


def test_the_readme_links_back_to_the_document(workbook):
    """Once both files are published the pair is navigable in both directions."""
    built, book = workbook
    sheet = book["Readme"]
    cell = next(c for row in sheet.iter_rows() for c in row if c.value == built["document"]["file"])
    assert cell.hyperlink.target == built["links"]["document"]


def test_the_readme_has_no_model_vocabulary(workbook):
    _built, book = workbook
    values = [str(cell.value) for row in book["Readme"].iter_rows() for cell in row if cell.value]
    assert not any("[]" in v or v.startswith("Vocabulaire") for v in values)
    assert L["workbook.readme.tabs_title"] in values


def test_an_endpoint_sheet_holds_the_request_above_the_response(workbook):
    _built, book = workbook
    sheet = book["assets"]
    request = block(sheet, L["workbook.blocks.request"])
    response = block(sheet, L["workbook.blocks.response"])
    assert request < response
    assert header(sheet, request) == list(L.section("workbook.request.columns").values())
    assert header(sheet, response) == list(L.section("workbook.response.columns").values())


def test_the_request_block_lists_every_parameter(workbook):
    built, book = workbook
    sheet = book["assets"]
    start = block(sheet, L["workbook.blocks.request"])
    params = [e for e in built["endpoints"] if e["name"] == "assets"][0]["params"]
    rows = [sheet.cell(row=start + 1 + i, column=1).value for i in range(len(params))]
    assert rows == [p["name"] for p in params]
    assert sheet.cell(row=start + 1, column=4).value == params[0]["origin"]


def test_an_endpoint_with_no_parameters_says_so(workbook):
    _built, book = workbook
    sheet = book["tenant_info"]
    start = block(sheet, L["workbook.blocks.request"])
    assert sheet.cell(row=start + 1, column=1).value == L["workbook.request.empty"]


def test_a_response_block_without_evidence_says_so(workbook):
    _built, book = workbook
    sheet = book["tenant_info"]
    start = block(sheet, L["workbook.blocks.response"])
    assert sheet.cell(row=start + 1, column=1).value == L["workbook.response.empty"]


def test_a_response_sheet_describes_the_body_as_observed(reference_source, reference_annotation, tmp_path):
    saved = [
        SavedOutput(
            Path("output/reference/assets/PUMP_p0.json"),
            build_envelope(reference_source, "assets", {"assetType": "PUMP"}, {"data": [{"id": {"id": "PUMP-1"}, "label": None}], "hasNext": True}),
        ),
        SavedOutput(
            Path("output/reference/assets/VALVE_p0.json"),
            build_envelope(reference_source, "assets", {"assetType": "VALVE"}, {"data": [], "hasNext": False}),
        ),
    ]
    built = model.build(reference_source, reference_annotation, ev.Evidence(envelopes={"assets": saved}))
    book = load_workbook(render_xlsx.render(built.model, tmp_path / "a.xlsx"))
    sheet = book["assets"]
    start = block(sheet, L["workbook.blocks.response"])
    rows = {
        sheet.cell(row=r, column=1).value: [sheet.cell(row=r, column=c).value for c in range(2, 6)]
        for r in range(start + 1, sheet.max_row + 1)
    }
    assert rows["$.hasNext"] == ["boolean", "Non", 1.0, "True"]
    assert rows["$.data[].id.id"][:2] == ["string", "Oui"]  # present in one response of two
    assert rows["$.data[].label"][:2] == ["null", "Oui"]
    assert sheet.cell(row=start + 1, column=4).number_format == "0 %"


def test_no_sheet_and_no_cell_mentions_volumes(workbook):
    _built, book = workbook
    for sheet in book.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                assert "volum" not in str(cell.value or "").lower(), (sheet.title, cell.coordinate)


@pytest.mark.parametrize("name, expected", [("assets", "assets"), ("a/b:c?", "a_b_c_"), ("x" * 40, "x" * 31)])
def test_sheet_titles_fit_excel(name, expected):
    assert model.sheet_title(name) == expected


def test_two_lists_named_alike_get_distinct_sheets():
    taken: set[str] = set()
    assert [model.unique_title("liste", taken) for _ in range(3)] == ["liste", "liste 2", "liste 3"]
